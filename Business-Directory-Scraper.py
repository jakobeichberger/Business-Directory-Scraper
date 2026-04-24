# -*- coding: utf-8 -*-
"""
Business Directory Scraper -> Excel / CSV

Scrapes paginated business-directory listing pages and their detail pages.
Detail pages are scanned as soon as their URLs are discovered, so progress is
written continuously instead of waiting until all listing pages were collected.
"""

from __future__ import annotations

import argparse
import csv
import html as html_lib
import json
import logging
import random
import re
import signal
import sys
import threading
import time
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Dict, Iterator, List, Optional, Set
from urllib.parse import parse_qs, unquote, urldefrag, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

try:
    import yaml

    _HAS_YAML = True
except ImportError:
    yaml = None
    _HAS_YAML = False


logger = logging.getLogger("scraper")
_thread_local = threading.local()
_search_rate_lock = threading.Lock()
_last_search_request_at = 0.0


# -----------------------------
# Defaults / constants
# -----------------------------
DEFAULT_BASE_URL = "https://www.firmenabc.at"
DEFAULT_START_URL = "https://www.firmenabc.at/firmen/wiener-neustadt-stadt_Fs"
DEFAULT_OUTPUT_XLSX = "business_listings.xlsx"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "de-AT,de;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}

DEFAULT_FIELD_SELECTORS = {
    "name": "h1",
    "email": "a[href^='mailto:']",
    "phone": "a[href^='tel:']",
    "website": "a[href^='http'], a[href*='url='], a[href*='website=']",
    "address": "",
}

XLSX_HEADERS = [
    "URL",
    "Name",
    "Address",
    "Phone",
    "Email",
    "Website",
    "UID",
    "Registry Number",
    "Credit Reference",
    "Scraped At",
    "Data Quality (%)",
    "Enriched",
]

QUALITY_FIELDS = ("name", "address", "phone", "email", "website")
LOW_QUALITY_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
HEADER_FONT = Font(bold=True)

TEXT_WEBSITE_PATTERN = re.compile(r"\b(?:https?://|www\.)[^\s<>()\"']+", re.IGNORECASE)
EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.IGNORECASE)
PHONE_PATTERN = re.compile(r"(?:\+\d{1,3}[\s./-]?)?(?:\(?\d+\)?[\s./-]?){6,}")

REDIRECT_URL_KEYS = (
    "uddg",
    "url",
    "u",
    "to",
    "target",
    "href",
    "link",
    "website",
    "redirect",
    "redirect_url",
)

SKIP_URL_EXTENSIONS = (
    ".7z",
    ".avi",
    ".bmp",
    ".css",
    ".csv",
    ".doc",
    ".docx",
    ".gif",
    ".gz",
    ".ics",
    ".jpeg",
    ".jpg",
    ".js",
    ".mov",
    ".mp3",
    ".mp4",
    ".pdf",
    ".png",
    ".rar",
    ".svg",
    ".txt",
    ".webp",
    ".xls",
    ".xlsx",
    ".zip",
)

SKIP_EXTERNAL_DOMAINS = (
    "bing.com",
    "duckduckgo.com",
    "facebook.com",
    "google.at",
    "google.com",
    "instagram.com",
    "linkedin.com",
    "maps.google.com",
    "tiktok.com",
    "twitter.com",
    "x.com",
    "xing.com",
    "youtube.com",
    "youtu.be",
)

WEBSITE_LINK_HINTS = (
    "homepage",
    "internet",
    "official",
    "web",
    "website",
    "webseite",
    "www.",
    "zur website",
)

CONTACT_LINK_HINTS = (
    "about",
    "anfahrt",
    "contact",
    "impressum",
    "imprint",
    "kontakt",
    "location",
    "standort",
    "team",
    "ueber",
    "uber",
    "über",
)

DEFAULT_WEBSITE_SCAN_PATHS = (
    "/kontakt",
    "/kontakt/",
    "/impressum",
    "/impressum/",
    "/contact",
    "/contact/",
    "/team",
    "/team/",
    "/ueber-uns",
    "/ueber-uns/",
    "/about",
    "/about/",
)

LEGAL_NAME_STOPWORDS = {
    "ag",
    "co",
    "das",
    "der",
    "die",
    "eu",
    "e.u",
    "gmbh",
    "kg",
    "mbh",
    "og",
    "und",
}


# -----------------------------
# Data model / config
# -----------------------------
@dataclass
class Config:
    base_url: str = DEFAULT_BASE_URL
    start_url: str = DEFAULT_START_URL
    output_xlsx: str = DEFAULT_OUTPUT_XLSX
    output_csv: str = ""
    http_threads: int = 1
    request_timeout: int = 25
    max_retries: int = 6
    max_listing_pages: int = 500
    playwright_wait_ms: int = 800
    listing_timeout_ms: int = 30000
    sleep_detail_min: float = 0.8
    sleep_detail_max: float = 2.0
    save_every: int = 1
    empty_listing_page_limit: int = 2
    repeated_listing_page_limit: int = 2
    detail_page_pattern: str = r"^/[^/?#]+_[A-Za-z0-9]+$"
    field_selectors: Dict[str, str] = field(default_factory=lambda: dict(DEFAULT_FIELD_SELECTORS))
    enrich_missing_data: bool = False
    search_fallback_enabled: bool = True
    search_fallback_when_quality_below: int = 80
    search_engines: str = "duckduckgo,bing"
    search_max_results: int = 5
    search_parallel_workers: int = 2
    search_request_delay: float = 0.5
    fallback_site_scan_limit: int = 3
    fallback_site_threads: int = 2
    website_scan_max_pages: int = 5
    website_scan_threads: int = 2
    website_scan_paths: List[str] = field(default_factory=lambda: list(DEFAULT_WEBSITE_SCAN_PATHS))
    headers: Dict[str, str] = field(default_factory=lambda: dict(DEFAULT_HEADERS))
    log_level: str = "INFO"
    log_file: str = ""


@dataclass
class ListingEntry:
    url: str
    name: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""
    website: str = ""
    uid: str = ""
    registry_number: str = ""
    credit_reference: str = ""
    scraped_at: str = ""
    enriched: bool = False

    @property
    def data_quality(self) -> int:
        filled = sum(1 for field_name in QUALITY_FIELDS if getattr(self, field_name))
        return round(filled / len(QUALITY_FIELDS) * 100)


@dataclass
class ScrapeStats:
    discovered: int = 0
    skipped: int = 0
    queued: int = 0
    written: int = 0
    failed: int = 0


@dataclass
class SearchResult:
    url: str
    title: str = ""
    snippet: str = ""
    source: str = ""


# -----------------------------
# Config / logging
# -----------------------------
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape business-directory listings to Excel.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--config", default="config.yaml", help="Path to YAML config file")
    parser.add_argument(
        "--log-level",
        default=None,
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Override log level from config",
    )
    return parser.parse_args()


def setup_logging(level: str = "INFO", log_file: str = "") -> None:
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    logger.handlers.clear()
    logger.propagate = False

    formatter = logging.Formatter(
        "[%(levelname)-7s] %(asctime)s  %(message)s",
        datefmt="%H:%M:%S",
    )

    console = logging.StreamHandler()
    console.setFormatter(formatter)
    logger.addHandler(console)

    if log_file:
        file_handler = logging.FileHandler(log_file, encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)


def _coerce_int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _coerce_float(value, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _coerce_bool(value, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on"}
    return default


def _strip_yaml_comment(line: str) -> str:
    quote: Optional[str] = None
    escaped = False

    for idx, char in enumerate(line):
        if escaped:
            escaped = False
            continue
        if char == "\\":
            escaped = True
            continue
        if char in {"'", '"'}:
            quote = None if quote == char else char if quote is None else quote
            continue
        if char == "#" and quote is None:
            return line[:idx].rstrip()

    return line.rstrip()


def _parse_simple_yaml_value(value: str):
    value = value.strip()
    if not value:
        return ""
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    if value.lower() in {"true", "false"}:
        return value.lower() == "true"
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def _load_simple_yaml(path: str) -> dict:
    """Fallback parser for this repo's simple config.yaml shape."""
    data: dict = {}
    current_section: Optional[str] = None

    with open(path, "r", encoding="utf-8") as fh:
        for raw_line in fh:
            line = _strip_yaml_comment(raw_line.rstrip("\n"))
            if not line.strip():
                continue

            indent = len(line) - len(line.lstrip(" "))
            stripped = line.strip()
            if ":" not in stripped:
                continue

            key, value = stripped.split(":", 1)
            key = key.strip()
            value = value.strip()

            if indent == 0:
                if value == "":
                    data[key] = {}
                    current_section = key
                else:
                    data[key] = _parse_simple_yaml_value(value)
                    current_section = None
            elif current_section and isinstance(data.get(current_section), dict):
                data[current_section][key] = _parse_simple_yaml_value(value)

    return data


def load_config(path: str) -> Config:
    cfg = Config()

    try:
        if _HAS_YAML:
            with open(path, "r", encoding="utf-8") as fh:
                data = yaml.safe_load(fh) or {}
        else:
            data = _load_simple_yaml(path)
    except FileNotFoundError:
        logger.warning("Config file not found: %s. Using built-in defaults.", path)
        return cfg

    cfg.base_url = str(data.get("base_url") or cfg.base_url)
    cfg.start_url = str(data.get("start_url") or cfg.start_url)
    cfg.output_xlsx = str(data.get("output_xlsx") or cfg.output_xlsx)
    cfg.output_csv = str(data.get("output_csv") or "")

    cfg.http_threads = max(1, _coerce_int(data.get("http_threads"), cfg.http_threads))
    cfg.request_timeout = max(1, _coerce_int(data.get("request_timeout"), cfg.request_timeout))
    cfg.max_retries = max(1, _coerce_int(data.get("max_retries"), cfg.max_retries))
    cfg.max_listing_pages = max(1, _coerce_int(data.get("max_listing_pages"), cfg.max_listing_pages))
    cfg.playwright_wait_ms = max(0, _coerce_int(data.get("playwright_wait_ms"), cfg.playwright_wait_ms))
    cfg.listing_timeout_ms = max(1000, _coerce_int(data.get("listing_timeout_ms"), cfg.listing_timeout_ms))
    cfg.save_every = max(1, _coerce_int(data.get("save_every"), cfg.save_every))
    cfg.empty_listing_page_limit = max(
        1,
        _coerce_int(data.get("empty_listing_page_limit"), cfg.empty_listing_page_limit),
    )
    cfg.repeated_listing_page_limit = max(
        1,
        _coerce_int(data.get("repeated_listing_page_limit"), cfg.repeated_listing_page_limit),
    )

    cfg.sleep_detail_min = max(0.0, _coerce_float(data.get("sleep_detail_min"), cfg.sleep_detail_min))
    cfg.sleep_detail_max = max(
        cfg.sleep_detail_min,
        _coerce_float(data.get("sleep_detail_max"), cfg.sleep_detail_max),
    )

    cfg.detail_page_pattern = str(data.get("detail_page_pattern") or cfg.detail_page_pattern)
    cfg.enrich_missing_data = _coerce_bool(data.get("enrich_missing_data"), cfg.enrich_missing_data)
    cfg.search_fallback_enabled = _coerce_bool(
        data.get("search_fallback_enabled"),
        cfg.search_fallback_enabled,
    )
    cfg.search_fallback_when_quality_below = max(
        0,
        min(100, _coerce_int(
            data.get("search_fallback_when_quality_below"),
            cfg.search_fallback_when_quality_below,
        )),
    )
    cfg.search_engines = str(data.get("search_engines") or cfg.search_engines)
    cfg.search_max_results = max(0, _coerce_int(data.get("search_max_results"), cfg.search_max_results))
    cfg.search_parallel_workers = max(
        1,
        _coerce_int(data.get("search_parallel_workers"), cfg.search_parallel_workers),
    )
    cfg.search_request_delay = max(
        0.0,
        _coerce_float(data.get("search_request_delay"), cfg.search_request_delay),
    )
    cfg.fallback_site_scan_limit = max(
        0,
        _coerce_int(data.get("fallback_site_scan_limit"), cfg.fallback_site_scan_limit),
    )
    cfg.fallback_site_threads = max(
        1,
        _coerce_int(data.get("fallback_site_threads"), cfg.fallback_site_threads),
    )
    cfg.website_scan_max_pages = max(
        0,
        _coerce_int(data.get("website_scan_max_pages"), cfg.website_scan_max_pages),
    )
    cfg.website_scan_threads = max(
        1,
        _coerce_int(data.get("website_scan_threads"), cfg.website_scan_threads),
    )
    if data.get("website_scan_paths"):
        cfg.website_scan_paths = [
            path.strip()
            for path in str(data["website_scan_paths"]).split(",")
            if path.strip()
        ]
    cfg.log_level = str(data.get("log_level") or cfg.log_level).upper()
    cfg.log_file = str(data.get("log_file") or "")

    selectors = dict(DEFAULT_FIELD_SELECTORS)
    selectors.update({k: str(v or "") for k, v in (data.get("field_selectors") or {}).items()})
    cfg.field_selectors = selectors

    headers = dict(DEFAULT_HEADERS)
    headers.update({str(k): str(v) for k, v in (data.get("headers") or {}).items() if v is not None})
    cfg.headers = headers

    return cfg


# -----------------------------
# Utility helpers
# -----------------------------
def _sleep(min_s: float, max_s: float) -> None:
    time.sleep(random.uniform(min_s, max_s))


def make_soup(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html, "lxml")
    except Exception:
        return BeautifulSoup(html, "html.parser")


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def _host(url: str) -> str:
    host = (urlparse(url).hostname or "").lower().rstrip(".")
    return host[4:] if host.startswith("www.") else host


def _is_same_or_subdomain(host: str, base_host: str) -> bool:
    return host == base_host or host.endswith(f".{base_host}")


def _is_same_site_url(url: str, site_url: str) -> bool:
    host = _host(url)
    site_host = _host(site_url)
    return bool(host and site_host and _is_same_or_subdomain(host, site_host))


def normalize_url(raw_url: str, base_url: str) -> str:
    raw = html_lib.unescape(_clean(raw_url))
    if not raw or raw.startswith("#"):
        return ""

    lower = raw.lower()
    if lower.startswith(("data:", "fax:", "javascript:", "mailto:", "sms:", "tel:")):
        return ""
    if lower.startswith("www."):
        raw = f"https://{raw}"

    absolute = urljoin(base_url, raw)
    absolute, _fragment = urldefrag(absolute)
    parsed = urlparse(absolute)

    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""

    return urlunparse((
        parsed.scheme.lower(),
        parsed.netloc.lower(),
        parsed.path or "/",
        "",
        parsed.query,
        "",
    ))


def external_url_from_href(href: str, page_url: str, cfg: Config) -> str:
    direct = normalize_url(href, page_url)
    if not direct:
        return ""

    query = parse_qs(urlparse(direct).query)
    for key in REDIRECT_URL_KEYS:
        for value in query.get(key, []):
            candidate = normalize_url(unquote(value), page_url)
            if candidate and not _is_same_site_url(candidate, cfg.base_url):
                return candidate

    return direct


def detail_url_from_href(href: str, cfg: Config, pattern: re.Pattern[str]) -> str:
    absolute = normalize_url(href, cfg.base_url)
    if not absolute or not _is_same_site_url(absolute, cfg.base_url):
        return ""

    parsed = urlparse(absolute)
    path = parsed.path.rstrip("/")
    if not pattern.match(path):
        return ""

    return urlunparse((parsed.scheme, parsed.netloc, path, "", "", ""))


def _normalize_email(raw: str) -> str:
    text = unquote(html_lib.unescape(_clean(raw))).replace("mailto:", "")
    text = text.split("?", 1)[0]
    match = EMAIL_PATTERN.search(text)
    return match.group(0).lower() if match else ""


def _normalize_phone(raw: str) -> str:
    text = unquote(html_lib.unescape(_clean(raw))).replace("tel:", "")
    text = re.sub(r"[^\d+]", "", text)
    if text.count("+") > 1:
        text = "+" + text.replace("+", "")
    return text


def _extract_email_from_text(text: str) -> str:
    match = EMAIL_PATTERN.search(text or "")
    return match.group(0).lower() if match else ""


def _extract_phone_from_text(text: str) -> str:
    for match in PHONE_PATTERN.finditer(text or ""):
        phone = _normalize_phone(match.group(0))
        digit_count = sum(1 for char in phone if char.isdigit())
        if 6 <= digit_count <= 20:
            return phone
    return ""


def name_from_detail_url(url: str) -> str:
    path = urlparse(url).path.rstrip("/")
    slug = path.rsplit("/", 1)[-1].split("_", 1)[0]
    slug = unquote(slug).replace("-", " ").replace("_", " ")
    return _clean(slug).title()


def _root_url(url: str) -> str:
    parsed = urlparse(url)
    if not parsed.scheme or not parsed.netloc:
        return ""
    return urlunparse((parsed.scheme, parsed.netloc, "/", "", "", ""))


def _dedupe_url_key(url: str) -> str:
    parsed = urlparse(url)
    path = parsed.path.rstrip("/") or "/"
    return urlunparse((parsed.scheme, parsed.netloc, path, "", parsed.query, ""))


def _name_tokens(name: str) -> Set[str]:
    raw_tokens = re.findall(r"[a-zA-ZÄÖÜäöüß0-9]{3,}", (name or "").lower())
    return {token for token in raw_tokens if token.replace(".", "") not in LEGAL_NAME_STOPWORDS}


def _result_mentions_name(result: SearchResult, name: str) -> bool:
    tokens = _name_tokens(name)
    if not tokens:
        return True

    haystack = f"{result.title} {result.snippet} {_host(result.url)} {urlparse(result.url).path}".lower()
    return any(token in haystack for token in tokens)


def merge_missing_fields(target: ListingEntry, source: ListingEntry) -> bool:
    changed = False
    for field_name in ("name", "address", "phone", "email", "website"):
        if not getattr(target, field_name) and getattr(source, field_name):
            setattr(target, field_name, getattr(source, field_name))
            changed = True

    if changed:
        target.enriched = True

    return changed


def score_contact_entry(entry: ListingEntry) -> int:
    return (
        (40 if entry.email else 0)
        + (35 if entry.phone else 0)
        + (20 if entry.address else 0)
        + (5 if entry.website else 0)
    )


def merge_best_fields(target: ListingEntry, candidates: List[ListingEntry]) -> bool:
    changed = False

    for candidate in sorted(candidates, key=score_contact_entry, reverse=True):
        changed = merge_missing_fields(target, candidate) or changed

    return changed


def contact_entry_from_text(text: str, url: str = "", name_hint: str = "") -> ListingEntry:
    return ListingEntry(
        url=url,
        name=name_hint,
        phone=_extract_phone_from_text(text),
        email=_extract_email_from_text(text),
        scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


def _is_skipped_external_host(host: str) -> bool:
    return any(_is_same_or_subdomain(host, skipped) for skipped in SKIP_EXTERNAL_DOMAINS)


def _looks_like_company_website(url: str, cfg: Config) -> bool:
    parsed = urlparse(url)
    host = _host(url)
    path = parsed.path.lower()

    if parsed.scheme.lower() not in {"http", "https"} or not host:
        return False
    if _is_same_site_url(url, cfg.base_url):
        return False
    if path.endswith(SKIP_URL_EXTENSIONS):
        return False
    if _is_skipped_external_host(host):
        return False
    if "google." in host and ("maps" in host or path.startswith("/maps")):
        return False

    return True


def _score_company_website(url: str, label: str, priority: int = 0) -> int:
    parsed = urlparse(url)
    haystack = f"{label} {url}".lower()
    score = priority

    if any(hint in haystack for hint in WEBSITE_LINK_HINTS):
        score += 25
    if parsed.path in {"", "/"}:
        score += 5
    if parsed.query:
        score -= 2

    return score


def _iter_json_urls(value, trusted_url_field: bool = False) -> Iterator[str]:
    if isinstance(value, dict):
        for key, child in value.items():
            yield from _iter_json_urls(child, key.lower() in {"sameas", "url", "website"})
    elif isinstance(value, list):
        for child in value:
            yield from _iter_json_urls(child, trusted_url_field)
    elif (
        trusted_url_field
        and isinstance(value, str)
        and re.match(r"^(?:https?://|www\.)", value, re.IGNORECASE)
    ):
        yield value


def extract_company_website(soup: BeautifulSoup, page_url: str, cfg: Config) -> str:
    candidates: List[tuple[int, int, str]] = []
    seen: Set[str] = set()

    def add_candidate(raw_url: str, label: str = "", priority: int = 0) -> None:
        url = external_url_from_href(raw_url, page_url, cfg)
        if not url or not _looks_like_company_website(url, cfg) or url in seen:
            return

        seen.add(url)
        candidates.append((
            _score_company_website(url, label, priority),
            len(candidates),
            url,
        ))

    for script in soup.select("script[type='application/ld+json']"):
        raw_json = script.string or script.get_text("", strip=True)
        if not raw_json:
            continue
        try:
            data = json.loads(raw_json)
        except json.JSONDecodeError:
            continue
        for raw_url in _iter_json_urls(data):
            add_candidate(raw_url, "structured data", priority=40)

    website_selector = cfg.field_selectors.get("website") or ""
    if website_selector:
        for a in soup.select(website_selector):
            if a.has_attr("href"):
                label = " ".join(filter(None, [
                    a.get_text(" ", strip=True),
                    a.get("title", ""),
                    a.get("aria-label", ""),
                ]))
                add_candidate(a.get("href", ""), label, priority=20)

    for a in soup.select("a[href]"):
        label = " ".join(filter(None, [
            a.get_text(" ", strip=True),
            a.get("title", ""),
            a.get("aria-label", ""),
        ]))
        add_candidate(a.get("href", ""), label, priority=10)

    for match in TEXT_WEBSITE_PATTERN.finditer(soup.get_text(" ", strip=True)):
        add_candidate(match.group(0).rstrip(".,;:)]}"), "visible text")

    if not candidates:
        return ""

    return sorted(candidates, key=lambda item: (-item[0], item[1]))[0][2]


def extract_contact_entry_from_soup(
    soup: BeautifulSoup,
    page_url: str,
    cfg: Config,
    name_hint: str = "",
) -> ListingEntry:
    text = soup.get_text("\n", strip=True)

    name = name_hint
    if not name:
        name_el = soup.select_one(cfg.field_selectors.get("name") or "h1")
        if name_el:
            name = _clean(name_el.get_text(" ", strip=True))
    if not name and soup.title:
        name = _clean(soup.title.get_text(" ", strip=True))

    email = ""
    email_el = soup.select_one("a[href^='mailto:']")
    if email_el:
        email = _normalize_email(email_el.get("href", ""))
    if not email:
        email = _extract_email_from_text(text)

    phone = ""
    phone_el = soup.select_one("a[href^='tel:']")
    if phone_el:
        phone = _normalize_phone(phone_el.get("href", ""))
    if not phone:
        phone = _extract_phone_from_text(text)

    website = normalize_url(page_url, page_url)
    if not _looks_like_company_website(website, cfg):
        website = extract_company_website(soup, page_url, cfg)

    address = ""
    lines = [line for line in text.splitlines() if line.strip()]
    for idx in range(len(lines) - 1):
        if re.search(r"\d", lines[idx]) and re.match(r"^\d{4,5}\s+\S+", lines[idx + 1]):
            address = _clean(f"{lines[idx]}, {lines[idx + 1]}")
            break

    return ListingEntry(
        url=page_url,
        name=name,
        address=address,
        phone=phone,
        email=email,
        website=website,
        scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


def iter_contact_page_candidates(soup: BeautifulSoup, site_url: str, cfg: Config) -> Iterator[str]:
    root = _root_url(site_url)
    seen: Set[str] = set()

    def add(raw_url: str) -> Iterator[str]:
        candidate = normalize_url(raw_url, root or site_url)
        candidate_key = _dedupe_url_key(candidate)
        if not candidate or candidate_key in seen:
            return
        if not _is_same_site_url(candidate, site_url):
            return
        seen.add(candidate_key)
        yield candidate

    for candidate in add(root or site_url):
        yield candidate

    for path in cfg.website_scan_paths:
        path = path if path.startswith("/") else f"/{path}"
        for candidate in add(urljoin(root or site_url, path)):
            yield candidate

    for a in soup.select("a[href]"):
        label = " ".join(filter(None, [
            a.get_text(" ", strip=True),
            a.get("title", ""),
            a.get("aria-label", ""),
            a.get("href", ""),
        ])).lower()
        if not any(hint in label for hint in CONTACT_LINK_HINTS):
            continue
        for candidate in add(a.get("href", "")):
            yield candidate


def scan_company_site(site_url: str, cfg: Config, name_hint: str = "") -> Optional[ListingEntry]:
    site_url = normalize_url(site_url, cfg.base_url)
    if not site_url or not _looks_like_company_website(site_url, cfg):
        return None

    site_root = _root_url(site_url) or site_url
    homepage_html = fetch_html(site_root, cfg)
    if not homepage_html:
        return ListingEntry(
            url=site_root,
            name=name_hint,
            website=site_root,
            scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        )

    homepage_soup = make_soup(homepage_html)
    combined = extract_contact_entry_from_soup(homepage_soup, site_root, cfg, name_hint)
    if cfg.website_scan_max_pages <= 1:
        combined.website = site_root or combined.website
        return combined

    candidate_urls = [
        candidate_url
        for candidate_url in iter_contact_page_candidates(homepage_soup, site_root, cfg)
        if candidate_url != site_root
    ][: max(0, cfg.website_scan_max_pages - 1)]

    candidate_entries: List[ListingEntry] = []
    max_workers = min(cfg.website_scan_threads, len(candidate_urls))

    def scan_contact_page(candidate_url: str) -> Optional[ListingEntry]:
        html = fetch_html(candidate_url, cfg)
        if not html:
            return None
        return extract_contact_entry_from_soup(make_soup(html), candidate_url, cfg, name_hint)

    if max_workers:
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            future_to_url = {pool.submit(scan_contact_page, url): url for url in candidate_urls}
            for future in as_completed(future_to_url):
                try:
                    candidate_entry = future.result()
                except Exception as exc:
                    logger.debug("Contact page scan failed for %s: %s", future_to_url[future], exc)
                    continue
                if candidate_entry:
                    candidate_entries.append(candidate_entry)

    merge_best_fields(combined, candidate_entries)

    combined.website = site_root or combined.website
    return combined


# -----------------------------
# HTTP
# -----------------------------
def _get_session(headers: Dict[str, str]) -> requests.Session:
    if not hasattr(_thread_local, "session"):
        session = requests.Session()
        session.headers.update(headers)
        _thread_local.session = session
    return _thread_local.session


def _retry_wait(attempt: int, cap: int = 30) -> float:
    return min(2 ** attempt, cap) + random.uniform(0, 0.5)


def fetch_html(url: str, cfg: Config) -> Optional[str]:
    session = _get_session(cfg.headers)
    last: object = None

    for attempt in range(1, cfg.max_retries + 1):
        try:
            response = session.get(url, timeout=cfg.request_timeout)

            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                wait = float(retry_after) if retry_after and retry_after.isdigit() else _retry_wait(attempt)
                last = f"HTTP 429 (waited {wait:.1f}s)"
                time.sleep(wait)
                continue

            if response.status_code in {408, 425} or response.status_code >= 500:
                wait = _retry_wait(attempt)
                last = f"HTTP {response.status_code} (waited {wait:.1f}s)"
                time.sleep(wait)
                continue

            response.raise_for_status()
            return response.text

        except requests.RequestException as exc:
            last = exc
            wait = _retry_wait(attempt, cap=20)
            logger.debug("Attempt %d/%d failed for %s: %s", attempt, cfg.max_retries, url, exc)
            time.sleep(wait)

    logger.warning("Failed permanently: %s -> %s", url, last)
    return None


def wait_for_search_slot(cfg: Config) -> None:
    global _last_search_request_at

    if cfg.search_request_delay <= 0:
        return

    with _search_rate_lock:
        now = time.monotonic()
        wait = (_last_search_request_at + cfg.search_request_delay) - now
        if wait > 0:
            time.sleep(wait)
        _last_search_request_at = time.monotonic()


def _fetch_search_page(url: str, params: Dict[str, str], cfg: Config) -> Optional[str]:
    session = _get_session(cfg.headers)
    try:
        wait_for_search_slot(cfg)
        response = session.get(url, params=params, timeout=cfg.request_timeout)
        response.raise_for_status()
        return response.text
    except requests.RequestException as exc:
        logger.debug("Search request failed for %s: %s", url, exc)
        return None


def search_duckduckgo_html(query: str, cfg: Config) -> List[SearchResult]:
    html = _fetch_search_page("https://duckduckgo.com/html/", {"q": query}, cfg)
    if not html:
        return []

    soup = make_soup(html)
    results: List[SearchResult] = []

    for result in soup.select(".result"):
        link = result.select_one("a.result__a[href]")
        if not link:
            continue

        url = external_url_from_href(link.get("href", ""), "https://duckduckgo.com/", cfg)
        if not url:
            continue

        snippet_el = result.select_one(".result__snippet")
        results.append(SearchResult(
            url=url,
            title=_clean(link.get_text(" ", strip=True)),
            snippet=_clean(snippet_el.get_text(" ", strip=True)) if snippet_el else "",
            source="duckduckgo",
        ))

    return results


def search_bing_html(query: str, cfg: Config) -> List[SearchResult]:
    html = _fetch_search_page("https://www.bing.com/search", {"q": query}, cfg)
    if not html:
        return []

    soup = make_soup(html)
    results: List[SearchResult] = []

    for result in soup.select("li.b_algo"):
        link = result.select_one("h2 a[href]")
        if not link:
            continue

        url = external_url_from_href(link.get("href", ""), "https://www.bing.com/", cfg)
        if not url:
            continue

        snippet_el = result.select_one("p")
        results.append(SearchResult(
            url=url,
            title=_clean(link.get_text(" ", strip=True)),
            snippet=_clean(snippet_el.get_text(" ", strip=True)) if snippet_el else "",
            source="bing",
        ))

    return results


def search_open_web(query: str, cfg: Config) -> List[SearchResult]:
    results: List[SearchResult] = []
    seen: Set[str] = set()
    engines = [engine.strip().lower() for engine in cfg.search_engines.split(",") if engine.strip()]

    def run_engine(engine: str) -> List[SearchResult]:
        if engine == "duckduckgo":
            return search_duckduckgo_html(query, cfg)
        if engine == "bing":
            return search_bing_html(query, cfg)
        logger.debug("Unknown search engine skipped: %s", engine)
        return []

    max_workers = min(max(1, cfg.search_parallel_workers), len(engines))
    if not max_workers:
        return []

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        future_to_engine = {pool.submit(run_engine, engine): engine for engine in engines}
        for future in as_completed(future_to_engine):
            engine = future_to_engine[future]
            try:
                engine_results = future.result()
            except Exception as exc:
                logger.debug("Search engine failed for %s: %s", engine, exc)
                continue

            for result in engine_results:
                normalized_url = normalize_url(result.url, cfg.base_url)
                normalized_key = _dedupe_url_key(normalized_url)
                if not normalized_url or normalized_key in seen:
                    continue
                seen.add(normalized_key)
                result.url = normalized_url
                results.append(result)

    return results[: cfg.search_max_results]


def build_search_queries(entry: ListingEntry) -> List[str]:
    name = entry.name or name_from_detail_url(entry.url)
    address = entry.address
    queries = []

    if name and address:
        queries.append(f'"{name}" "{address}" kontakt telefon email website')
    if name:
        queries.append(f'"{name}" kontakt telefon email website')
        queries.append(f'"{name}" impressum')
    if entry.url:
        queries.append(f'"{name_from_detail_url(entry.url)}" kontakt')

    return list(dict.fromkeys(query for query in queries if query.strip()))


def search_fallback_for_entry(entry: ListingEntry, cfg: Config, force: bool = False) -> ListingEntry:
    if not cfg.search_fallback_enabled:
        return entry
    if not force and entry.data_quality >= cfg.search_fallback_when_quality_below:
        return entry

    name_hint = entry.name or name_from_detail_url(entry.url)
    logger.info("[SEARCH] Fallback for: %s", name_hint or entry.url)

    for query in build_search_queries(entry):
        results = search_open_web(query, cfg)
        logger.debug("[SEARCH] %d results for %s", len(results), query)
        matching_results = [result for result in results if _result_mentions_name(result, name_hint)]

        snippet_entries: List[ListingEntry] = []
        site_results: List[SearchResult] = []
        for result in matching_results:
            text_entry = contact_entry_from_text(
                f"{result.title} {result.snippet}",
                url=entry.url,
                name_hint=name_hint,
            )
            snippet_entries.append(text_entry)

            if _looks_like_company_website(result.url, cfg):
                site_results.append(result)

        merge_best_fields(entry, snippet_entries)

        site_results = site_results[: cfg.fallback_site_scan_limit]
        site_entries: List[ListingEntry] = []
        max_workers = min(cfg.fallback_site_threads, len(site_results))

        if max_workers:
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                future_to_result = {
                    pool.submit(scan_company_site, result.url, cfg, name_hint): result
                    for result in site_results
                }
                for future in as_completed(future_to_result):
                    result = future_to_result[future]
                    try:
                        site_entry = future.result()
                    except Exception as exc:
                        logger.debug("Fallback site scan failed for %s: %s", result.url, exc)
                        continue
                    if site_entry:
                        site_entries.append(site_entry)

        merge_best_fields(entry, site_entries)

        if entry.data_quality >= cfg.search_fallback_when_quality_below and entry.phone and entry.email:
            return entry

    return entry


# -----------------------------
# Listing pages
# -----------------------------
def _listing_page_url(start_url: str, page_number: int) -> str:
    return start_url if page_number == 1 else f"{start_url.rstrip('/')}/{page_number}"


def load_listing_page_html(page, url: str, cfg: Config) -> Optional[str]:
    last: object = None

    for attempt in range(1, cfg.max_retries + 1):
        try:
            response = page.goto(url, wait_until="domcontentloaded")
            page.wait_for_timeout(cfg.playwright_wait_ms)

            if response and (response.status in {408, 425, 429} or response.status >= 500):
                raise RuntimeError(f"HTTP {response.status}")

            return page.content()

        except Exception as exc:
            last = exc
            wait = _retry_wait(attempt, cap=20)
            logger.warning("[LIST] Attempt %d/%d failed for %s: %s", attempt, cfg.max_retries, url, exc)
            time.sleep(wait)

    logger.warning("[LIST] Giving up on %s: %s", url, last)
    return None


def iter_listing_links_playwright(cfg: Config) -> Iterator[str]:
    pattern = re.compile(cfg.detail_page_pattern, re.IGNORECASE)
    seen: Set[str] = set()

    headers = dict(cfg.headers)
    user_agent = headers.pop("User-Agent", None)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context_args = {"extra_http_headers": headers}
        if user_agent:
            context_args["user_agent"] = user_agent

        context = browser.new_context(**context_args)
        page = context.new_page()
        page.set_default_timeout(cfg.listing_timeout_ms)

        empty_pages = 0
        repeated_pages = 0

        try:
            for page_number in range(1, cfg.max_listing_pages + 1):
                url = _listing_page_url(cfg.start_url, page_number)
                logger.info("[LIST] Page %d loading: %s", page_number, url)

                html = load_listing_page_html(page, url, cfg)
                if not html:
                    break

                soup = make_soup(html)
                page_links: List[str] = []
                for a in soup.select("a[href]"):
                    detail_url = detail_url_from_href(a.get("href", ""), cfg, pattern)
                    if detail_url:
                        page_links.append(detail_url)

                page_links = list(dict.fromkeys(page_links))

                if not page_links:
                    empty_pages += 1
                    logger.info(
                        "[LIST] Page %d had no detail links (%d/%d)",
                        page_number,
                        empty_pages,
                        cfg.empty_listing_page_limit,
                    )
                    if empty_pages >= cfg.empty_listing_page_limit:
                        break
                    continue

                empty_pages = 0
                new_links = 0

                for detail_url in page_links:
                    if detail_url in seen:
                        continue
                    seen.add(detail_url)
                    new_links += 1
                    yield detail_url

                if new_links == 0:
                    repeated_pages += 1
                    logger.info(
                        "[LIST] Page %d repeated known links (%d/%d)",
                        page_number,
                        repeated_pages,
                        cfg.repeated_listing_page_limit,
                    )
                    if repeated_pages >= cfg.repeated_listing_page_limit:
                        break
                else:
                    repeated_pages = 0

                logger.info("[LIST] Page %d: +%d new links (total: %d)", page_number, new_links, len(seen))

        finally:
            context.close()
            browser.close()


# -----------------------------
# Detail parser / enrichment
# -----------------------------
def parse_listing_html(url: str, html: str, cfg: Config) -> ListingEntry:
    soup = make_soup(html)
    selectors = cfg.field_selectors

    name_el = soup.select_one(selectors.get("name") or "h1")
    name = _clean(name_el.get_text(" ", strip=True)) if name_el else ""

    email = ""
    email_sel = selectors.get("email") or "a[href^='mailto:']"
    email_el = soup.select_one(email_sel)
    if email_el:
        email = _normalize_email(email_el.get("href", "") or email_el.get_text(" ", strip=True))
    if not email:
        email_el = soup.select_one("a[href^='mailto:']")
        if email_el:
            email = _normalize_email(email_el.get("href", ""))

    phone = ""
    phone_sel = selectors.get("phone") or "a[href^='tel:']"
    phone_el = soup.select_one(phone_sel)
    if phone_el:
        phone = _normalize_phone(phone_el.get("href", "") or phone_el.get_text(" ", strip=True))
    if not phone:
        phone_el = soup.select_one("a[href^='tel:']")
        if phone_el:
            phone = _normalize_phone(phone_el.get("href", ""))

    website = extract_company_website(soup, url, cfg)

    address = ""
    address_sel = selectors.get("address") or ""
    if address_sel:
        address_el = soup.select_one(address_sel)
        if address_el:
            address = _clean(address_el.get_text(" ", strip=True))

    if not address:
        lines = [line for line in soup.get_text("\n", strip=True).splitlines() if line.strip()]
        for idx in range(len(lines) - 1):
            if re.search(r"\d", lines[idx]) and re.match(r"^\d{4,5}\s+\S+", lines[idx + 1]):
                address = _clean(f"{lines[idx]}, {lines[idx + 1]}")
                break

    return ListingEntry(
        url=url,
        name=name,
        address=address,
        phone=phone,
        email=email,
        website=website,
        scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    )


def enrich_missing_fields(entry: ListingEntry, cfg: Config) -> ListingEntry:
    if not cfg.enrich_missing_data:
        return entry
    if entry.phone and entry.email and entry.website:
        return entry
    if not entry.name:
        return entry

    query = " ".join(part for part in [entry.name, entry.address, "kontakt website"] if part)
    session = _get_session(cfg.headers)

    try:
        response = session.get(
            "https://api.duckduckgo.com/",
            params={"q": query, "format": "json", "no_html": 1, "skip_disambig": 1},
            timeout=cfg.request_timeout,
        )
        response.raise_for_status()
        data = response.json()
    except Exception as exc:
        logger.debug("Enrichment failed for %s: %s", entry.url, exc)
        return entry

    changed = False
    text_blob = json.dumps(data, ensure_ascii=False)

    if not entry.email:
        match = EMAIL_PATTERN.search(text_blob)
        if match:
            entry.email = match.group(0).lower()
            changed = True

    if not entry.phone:
        match = PHONE_PATTERN.search(text_blob)
        if match:
            entry.phone = _normalize_phone(match.group(0))
            changed = True

    if not entry.website:
        urls: List[str] = []
        for key in ("AbstractURL", "Redirect"):
            if data.get(key):
                urls.append(str(data[key]))
        for result in data.get("Results", []):
            if result.get("FirstURL"):
                urls.append(str(result["FirstURL"]))
        for raw_url in urls:
            candidate = normalize_url(raw_url, cfg.base_url)
            if candidate and _looks_like_company_website(candidate, cfg):
                entry.website = candidate
                changed = True
                break

    entry.enriched = changed
    return entry


def fetch_and_parse_listing(url: str, cfg: Config) -> Optional[ListingEntry]:
    try:
        html = fetch_html(url, cfg)
        if not html:
            entry = ListingEntry(
                url=url,
                name=name_from_detail_url(url),
                scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            )
            return search_fallback_for_entry(entry, cfg, force=True)

        entry = parse_listing_html(url, html, cfg)
        entry = enrich_missing_fields(entry, cfg)
        return search_fallback_for_entry(entry, cfg)
    except Exception as exc:
        logger.warning("Could not parse %s: %s", url, exc)
        entry = ListingEntry(
            url=url,
            name=name_from_detail_url(url),
            scraped_at=datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        )
        return search_fallback_for_entry(entry, cfg, force=True)
    finally:
        _sleep(cfg.sleep_detail_min, cfg.sleep_detail_max)


# -----------------------------
# Excel / CSV
# -----------------------------
def ensure_headers(ws) -> None:
    for col_idx, header in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT


def load_or_create_workbook(path: str):
    try:
        wb = load_workbook(path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    ensure_headers(ws)
    return wb, ws


def read_done_urls(ws) -> Set[str]:
    return {
        ws.cell(row=row_idx, column=1).value
        for row_idx in range(2, ws.max_row + 1)
        if isinstance(ws.cell(row=row_idx, column=1).value, str)
    }


def append_row(ws, row: ListingEntry) -> None:
    ws.append([
        row.url,
        row.name,
        row.address,
        row.phone,
        row.email,
        row.website,
        row.uid,
        row.registry_number,
        row.credit_reference,
        row.scraped_at,
        row.data_quality,
        "Yes" if row.enriched else "No",
    ])

    if row.data_quality < 60:
        for col_idx in range(1, len(XLSX_HEADERS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = LOW_QUALITY_FILL


def autosize_columns(ws, cap: int = 60) -> None:
    for col_idx, _header in enumerate(XLSX_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, cap)


def save_csv(entries: List[ListingEntry], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(XLSX_HEADERS)
        for entry in entries:
            writer.writerow([
                entry.url,
                entry.name,
                entry.address,
                entry.phone,
                entry.email,
                entry.website,
                entry.uid,
                entry.registry_number,
                entry.credit_reference,
                entry.scraped_at,
                entry.data_quality,
                "Yes" if entry.enriched else "No",
            ])

    logger.info("CSV saved: %s (%d rows from this run)", path, len(entries))


def save_workbook(wb, ws, path: str) -> None:
    autosize_columns(ws)
    wb.save(path)


# -----------------------------
# Streaming pipeline
# -----------------------------
def handle_finished_future(
    future: Future,
    futures: Dict[Future, str],
    wb,
    ws,
    lock: threading.Lock,
    cfg: Config,
    stats: ScrapeStats,
    entries: List[ListingEntry],
    done_urls: Set[str],
) -> None:
    url = futures.pop(future)

    try:
        row = future.result()
    except Exception as exc:
        stats.failed += 1
        logger.warning("[DETAIL] Failed: %s -> %s", url, exc)
        return

    if not row:
        stats.failed += 1
        logger.warning("[DETAIL] Empty result: %s", url)
        return

    with lock:
        if row.url in done_urls:
            stats.skipped += 1
            return

        append_row(ws, row)
        done_urls.add(row.url)
        entries.append(row)
        stats.written += 1

        if stats.written % cfg.save_every == 0:
            save_workbook(wb, ws, cfg.output_xlsx)

    logger.info(
        "[SAVE] %d saved: %s quality=%d%% pending=%d",
        stats.written,
        row.name or row.url,
        row.data_quality,
        len(futures),
    )


def drain_finished_futures(
    futures: Dict[Future, str],
    wb,
    ws,
    lock: threading.Lock,
    cfg: Config,
    stats: ScrapeStats,
    entries: List[ListingEntry],
    done_urls: Set[str],
    block: bool = False,
) -> None:
    if block:
        for future in as_completed(list(futures)):
            handle_finished_future(future, futures, wb, ws, lock, cfg, stats, entries, done_urls)
        return

    for future in list(futures):
        if future.done():
            handle_finished_future(future, futures, wb, ws, lock, cfg, stats, entries, done_urls)


def main() -> None:
    args = parse_args()

    setup_logging("INFO")
    cfg = load_config(args.config)
    if args.log_level:
        cfg.log_level = args.log_level

    setup_logging(cfg.log_level, cfg.log_file)

    logger.info("Config loaded from: %s", args.config)
    logger.info("Listing scan starts at: %s", cfg.start_url)
    logger.info("Output Excel: %s", cfg.output_xlsx)
    logger.info("Streaming detail scan: enabled, save_every=%d", cfg.save_every)

    wb, ws = load_or_create_workbook(cfg.output_xlsx)
    lock = threading.Lock()
    done_urls = read_done_urls(ws)

    logger.info("Resume: %d URLs already saved", len(done_urls))

    stats = ScrapeStats()
    entries: List[ListingEntry] = []
    scheduled_urls = set(done_urls)
    futures: Dict[Future, str] = {}
    interrupted = False

    try:
        with ThreadPoolExecutor(max_workers=cfg.http_threads) as pool:
            for detail_url in iter_listing_links_playwright(cfg):
                stats.discovered += 1

                if detail_url in scheduled_urls:
                    stats.skipped += 1
                    drain_finished_futures(futures, wb, ws, lock, cfg, stats, entries, done_urls)
                    continue

                scheduled_urls.add(detail_url)
                futures[pool.submit(fetch_and_parse_listing, detail_url, cfg)] = detail_url
                stats.queued += 1
                logger.info("[QUEUE] Detail queued: %s (pending=%d)", detail_url, len(futures))

                drain_finished_futures(futures, wb, ws, lock, cfg, stats, entries, done_urls)

            drain_finished_futures(futures, wb, ws, lock, cfg, stats, entries, done_urls, block=True)

    except KeyboardInterrupt:
        interrupted = True
        logger.warning("Interrupted by user. Saving progress...")

    with lock:
        save_workbook(wb, ws, cfg.output_xlsx)

    if cfg.output_csv:
        save_csv(entries, cfg.output_csv)

    status = "INTERRUPTED" if interrupted else "DONE"
    logger.info(
        "[%s] discovered=%d queued=%d written=%d skipped=%d failed=%d -> %s",
        status,
        stats.discovered,
        stats.queued,
        stats.written,
        stats.skipped,
        stats.failed,
        cfg.output_xlsx,
    )


if __name__ == "__main__":
    signal.signal(signal.SIGINT, signal.default_int_handler)
    try:
        main()
    except Exception as exc:
        logger.exception("Fatal error: %s", exc)
        sys.exit(1)
